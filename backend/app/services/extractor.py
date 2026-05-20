"""Lightweight text extraction for uploaded docs."""
import csv
import io
import json
from io import BytesIO
from pathlib import Path

from .filetype import classify


def extract_text(filename: str, data: bytes) -> str:
    ext = Path(filename).suffix.lower()
    if ext in (".txt", ".md", ".markdown", ".log"):
        return data.decode("utf-8", errors="replace")
    if ext == ".csv":
        return _csv_to_text(data, delimiter=",")
    if ext == ".tsv":
        return _csv_to_text(data, delimiter="\t")
    if ext in (".json", ".jsonl"):
        return _json_to_text(data, jsonl=(ext == ".jsonl"))
    if ext in (".xlsx", ".xls"):
        return _xlsx_to_text(data)
    if ext == ".pdf":
        from pypdf import PdfReader
        reader = PdfReader(BytesIO(data))
        return "\n\n".join((p.extract_text() or "") for p in reader.pages)
    if ext == ".docx":
        from docx import Document as Docx
        d = Docx(BytesIO(data))
        return "\n".join(p.text for p in d.paragraphs)
    # Fallback: best-effort decode
    return data.decode("utf-8", errors="replace")


def _csv_to_text(data: bytes, *, delimiter: str) -> str:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return ""
    header, *body = rows
    lines = []
    for r in body:
        parts = [f"{h}: {v}" for h, v in zip(header, r) if v]
        if parts:
            lines.append(" | ".join(parts))
    return "\n".join(lines) if lines else "\n".join(delimiter.join(r) for r in rows)


def _json_to_text(data: bytes, *, jsonl: bool) -> str:
    text = data.decode("utf-8", errors="replace")
    try:
        if jsonl:
            items = [json.loads(line) for line in text.splitlines() if line.strip()]
        else:
            parsed = json.loads(text)
            items = parsed if isinstance(parsed, list) else [parsed]
    except Exception:
        return text
    return "\n\n".join(json.dumps(it, ensure_ascii=False, indent=2) for it in items)


def _xlsx_to_text(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return ""
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        out.append(f"# Sheet: {ws.title}")
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c is not None else "" for c in rows[0]]
        for r in rows[1:]:
            parts = [f"{h}: {v}" for h, v in zip(header, r) if v not in (None, "")]
            if parts:
                out.append(" | ".join(parts))
    return "\n".join(out)


def chunk_text(text: str, *, chunk_size: int = 800, overlap: int = 120) -> list[str]:
    text = (text or "").strip()
    if not text:
        return []
    # Split by paragraphs first, then pack
    paras = [p.strip() for p in text.split("\n\n") if p.strip()]
    chunks: list[str] = []
    buf = ""
    for p in paras:
        if len(buf) + len(p) + 2 <= chunk_size:
            buf = f"{buf}\n\n{p}" if buf else p
        else:
            if buf:
                chunks.append(buf)
            if len(p) <= chunk_size:
                buf = p
            else:
                # hard-split long paragraph with overlap
                i = 0
                while i < len(p):
                    chunks.append(p[i:i + chunk_size])
                    i += chunk_size - overlap
                buf = ""
    if buf:
        chunks.append(buf)
    return chunks
